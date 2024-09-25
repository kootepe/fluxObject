#!/usr/bin/env python3

import os
import logging
from pathlib import Path

import openpyxl as opxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from matplotlib.figure import Figure
from matplotlib.backends.backend_agg import FigureCanvasAgg as FigureCanvas
import matplotlib.patches as patches

from matplotlib.dates import date2num

from tools.time_funcs import rm_tz

logger = logging.getLogger("defaultLogger")


def create_excel(df, path, sort=None, name=None):
    # excel can't handle timestamps with timezones
    df = rm_tz(df)
    # initiate worksheet
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Fluxes"
    ws1 = wb.worksheets[0]

    # sorting by column goes here
    if sort is not None:
        pass

    # file naming
    if name is None:
        file_date = df.index[0].strftime("%Y%m%d")
        xlsx_name = f"{path}/{file_date}.xlsx"
    else:
        xlsx_name = f"{path}/{name}.xlsx"
    logger.debug(f"Creating file {xlsx_name}.")

    # dataframe rows to excel rows
    logger.debug(f"Dataframe to excel rows")
    for idx, r in enumerate(dataframe_to_rows(df, index=False, header=True)):
        ws1.append(r)
        ws1.row_dimensions[idx + 1].height = 15

    # columns which are appended with "fig_dir_" have paths to figs
    logger.debug(f"{df.columns}")
    cols = [d for d in df.columns if "fig_dir_" in d]
    # each column with plot paths to list
    figs = [df[col].to_list() for col in cols]

    logger.info(f"Adding figs to xlsx.")
    for idxx, fig_ls in enumerate(figs):
        # add column for each different gas
        ws1.insert_cols(idxx + 1)
        # adjust column width for the the plot figure
        ws1.column_dimensions[get_column_letter(idxx + 1)].width = 13
        # first row is the column name
        ws1.cell(row=1, column=idxx + 1).value = f"{cols[idxx][-3:]}_graph"
        for row, fig in enumerate(fig_ls):
            anc_str = get_column_letter(idxx + 1) + str(row + 2)
            try:
                img = opxl.drawing.image.Image(fig)
            except Exception:
                logger.debug("No fig")
                continue
            ws1.add_image(img, anc_str)

    path = Path(path)
    if not path.exists():
        path.mkdir(parents=True)
    # NOTE: CHECK IF FILE EXISTS AND USE ANOTHER NAME IF IT DOES
    wb.save(xlsx_name)
    logger.info(f"Saved .xlsx in {xlsx_name}")


def create_fig():
    fig = Figure(figsize=(0.8, 0.175))
    ax = fig.add_subplot()
    ax.set_frame_on(False)

    return fig, ax


def create_sparkline(df, filename, gas, fig, ax, rects):
    """
    Create a sparkline (small graph) that's on each row of the output excel

    Parameters
    ----------
    df : pd.DataFrame
    filename : datetime.datetime
    """
    wrec_x, wrec_w, srec_x, srec_w, rec_y, rec_h = rects
    w_rect = patches.Rectangle(
        (wrec_x, rec_y), wrec_w, rec_h, ec="grey", fc="grey", alpha=0.2
    )
    s_rect = patches.Rectangle(
        (srec_x, rec_y), srec_w, rec_h, ec="green", fc="green", alpha=0.2
    )
    ax.plot(df, color="r", linewidth=0.3)
    ax.add_patch(w_rect)
    ax.add_patch(s_rect)
    ax.set_xticks([], [])
    ax.set_yticks([], [])
    canvas = FigureCanvas(fig)
    canvas.print_figure(filename, dpi=150, bbox_inches="tight", pad_inches=0.005)
    ax.cla()


def create_rects(y, times):
    """
    Create bounds for rectangles displayed in the gas measurement sparklines

    Parameters
    ----------
    y : pd.DataFrame
        The gas measurement
    times : filter_tuple
        Filter tuple with the chamber close time and open time
    m_times :
        Filter tuple where the gas flux is measured from


    Returns
    -------
    rects : tuple
        list of the bounds of the displayed rectangles
    """
    whole_measurement_x = date2num(times.plot_start)
    whole_measurement_width = date2num(times.plot_end) - date2num(times.plot_start)
    calc_x = date2num(times.close)
    calc_width = date2num(times.open) - date2num(times.close)
    rectangle_y = y.min()
    rectangle_h = y.max() - y.min()
    rects = (
        whole_measurement_x,
        whole_measurement_width,
        calc_x,
        calc_width,
        rectangle_y,
        rectangle_h,
    )
    return rects
